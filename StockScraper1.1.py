from lxml import html #import librarie
import requests #import librarie
import xlrd, xlwt #import librarie
from openpyxl import *
from tkinter import *
from tkinter import ttk

###ONLINE SEARCH FUNCTION###

def search(stockticker): # definiton of seach function
    headers = requests.utils.default_headers()
    headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:20.0) Gecko/20100101 Firefox/20.0',
    })
    print('Stock ticker: ', stockticker)
    url = ("https://finance.yahoo.com/quote/%s" % stockticker)
    print(url)
    urlopen = requests.get(url, headers = headers)
    tree = html.fromstring(urlopen.content, "lxml")
    global price
    price = tree.xpath('//span[@class="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)"]/text()')
    #description = tree.xpath('//p[@class="description__text"]/text()')
    global day_low
    global low_high
    day_low = tree.xpath('//span[@class="Trsdu(0.3s) "]/text()')
    low_high = tree.xpath('//td[@class="Ta(end) Fw(b) Lh(14px)"]/text()')
    #print('Description: ', description)
    print('Current stock price: $', price[0])
    print('Open stock price: $', day_low[1])
    print("Today's high and low prices: Low", low_high[0], 'High')
    if stockticker not in portfoliotickerlist:
        PortfolioBtn = Button(tab1, text='Add to Portfolio', command=searchaddtoportfolio(ticker))
        PortfolioBtn.grid(column=2,row=4)
        myportfolio()
    addtoportfoliobtn = Button(tab1, text='Buy Stock', command=buystock)
    addtoportfoliobtn.grid(column=2, row=5)
    lbl4 = Label(tab4, text='Your wallet is %s' % walletvalue)
    lbl4.grid(column=0, row=0)
    global low_high_split
    low_high_split = low_high[0].split()
    return price, day_low, low_high_split

###COMPANY SEARCH FUNCTION###

def gosearch(companyname):
    usersearch = companyname.title()
    global company_name
    company_name = usersearch
    print(usersearch)
    workbook = xlrd.open_workbook('CompanyFile.xlsx', on_demand=True)
    worksheet = workbook.sheet_by_index(0)
    i = 0
    g = 0
    counter = 0
    for i in range(worksheet.nrows):
        if usersearch in worksheet.cell(i,1).value:
            global ticker
            ticker = worksheet.cell_value(i,0)
            print("Your company is: ", worksheet.cell_value(i,1))
            search(ticker)
        else:
            counter += 1
        if counter == 3312:
            print("Company not found")

###TK CODE###

def clicked():
    #pulls up info
    gosearch(txt.get())
    lbl1.configure(text = (txt.get()).title())
    currentP.configure(text = price)
    OpenP.configure(text = day_low[1])
    DayLow.configure(text = low_high_split[0])
    DayHigh.configure(text= low_high_split[2])

###WATCHLIST###

def initialportfolio():
    global portfolio
    global portfoliovalue
    global portfoliotickerlist
    portfolio = []
    portfoliovalue = []
    portfoliotickerlist = []
    workbook = xlrd.open_workbook('CompanyFile.xlsx', on_demand=True)
    global row_count
    pworksheet = workbook.sheet_by_index(1)
    global row_count
    row_count = pworksheet.nrows
    if row_count != 0:
        for i in range(pworksheet.nrows):
            tickercell = pworksheet.cell(i, 0).value
            companycell = pworksheet.cell(i, 1).value
            cellinput = companycell, tickercell
            portfoliotickerlist.append(tickercell)
            portfolio.append(cellinput)
        for i in range(pworksheet.nrows-1):
            stockvalue = pworksheet.cell(i+1, 2).value
            if stockvalue != '':
                portfoliovalue.append(float(stockvalue))

def myportfolio():
    global portfolio
    global portfoliovalue
    global portfoliotickerlist
    portfolio = []
    portfoliovalue = []
    portfoliotickerlist = []
    workbook = xlrd.open_workbook('CompanyFile.xlsx', on_demand=True)
    global row_count
    pworksheet = workbook.sheet_by_index(1)
    global row_count
    row_count = pworksheet.nrows
    if row_count != 0:
        for i in range(pworksheet.nrows):
            tickercell = pworksheet.cell(i, 0).value
            companycell = pworksheet.cell(i, 1).value
            cellinput = companycell, tickercell
            portfoliotickerlist.append(tickercell)
            portfolio.append(cellinput)
        for i in range(pworksheet.nrows - 1):
            stockvalue = pworksheet.cell(i + 1, 2).value
            if stockvalue != '':
                portfoliovalue.append(float(stockvalue))
                print(portfoliovalue)
                print(sum(portfoliovalue))
    portfolioBox = Listbox(tab2)
    portfolioBox.grid(column=0, row=0)
    for item in portfolio:
        portfolioBox.insert(END, item)

def addtoportfolio():
    searchtype = int(input("To search by ticker select 1 \n To search by company name select 2"))
    if searchtype == 1:
        companysearch = input("What is the company's ticker?").upper()
        search(companysearch)
        #validate ticker
    elif searchtype == 2:
        companysearch = input("What is the name of the company?")
        gosearch(companysearch)
        #validate company\

def searchaddtoportfolio(stockticker):
    wb = load_workbook(filename='CompanyFile.xlsx')
    sheet_ranges = wb['Sheet2']
    i = row_count
    if i == 1:
        sheet_ranges.cell(2, 1).value = stockticker
        sheet_ranges.cell(2, 2).value = company_name
        wb.save('CompanyFile.xlsx')
    else:
        sheet_ranges.cell(i+1, 1).value = stockticker
        sheet_ranges.cell(i + 1, 2).value = company_name
        wb.save('CompanyFile.xlsx')

###WALLET CODE###

def setwalletvalue():
    global walletvalue
    try:
        wb = load_workbook(filename='CompanyFile.xlsx')
        sheet_ranges = wb['Sheet3']
        walletvalue = int(input("Enter your initial deposit amount."))
        sheet_ranges.cell(1, 1).value = walletvalue
        wb.save('CompanyFile.xlsx')
        initialwalletvalue()
    except:
        1 == 1
    return walletvalue

def initialwalletvalue():
    global walletvalue
    workbook = xlrd.open_workbook('CompanyFile.xlsx', on_demand=True)
    worksheet = workbook.sheet_by_index(2)
    if worksheet.nrows != 0:
        walletvalue = worksheet.cell_value(0, 0)
    else:
        wb = load_workbook(filename='CompanyFile.xlsx')
        sheet_ranges = wb['Sheet3']
        walletvalue = int(input("Enter your initial deposit amount."))
        sheet_ranges.cell(1, 1).value = walletvalue
        wb.save('CompanyFile.xlsx')
    lbl4 = Label(tab4, text='Your wallet is $%.2f' % walletvalue)
    lbl4.grid(column=0, row=0)
    lbl5 = Label(tab4, text='Your portfolio value is $%.2f' % sum(portfoliovalue))
    lbl5.grid(column=3, row=0)

def editwalletvalue(cost,walletvalue):
    wb = load_workbook(filename='CompanyFile.xlsx')
    sheet_ranges = wb['Sheet3']
    walletvalue = walletvalue - cost
    sheet_ranges.cell(1, 1).value = walletvalue
    wb.save('CompanyFile.xlsx')

def buystock():
    count = int(input("How much stock do you want to buy?"))
    buystockammountcheck(count)

def buystockammountcheck(count):
    global stockprice
    stockprice = float(price[0].replace(",",""))
    for number in range(count+1):
        if number*stockprice < walletvalue:
            global newcount
            newcount = number
    cost = newcount*stockprice
    if newcount == 0:
        print("You have insufficient funds to buy 1 share of", company_name+".")
    elif newcount > 0:
        buypromt(count,newcount,cost, walletvalue)

def buypromt(count,newcount,cost, walletvalue):
    if count == newcount:
        print("Press 1 to purchase", count, "shares of", company_name + ".")
        userchoice = int(input())
    elif count != newcount:
        print("You can only purchase", newcount, "share(s) of", company_name+".")
        print("Press 1 to purchase", newcount, "share(s) of", company_name + ".")
        userchoice = int(input())
    if userchoice == 1:
        editwalletvalue(cost,walletvalue)
        wb = load_workbook(filename='CompanyFile.xlsx')
        sheet_ranges = wb['Sheet2']
        i = row_count
    if i == 1:
        sheet_ranges.cell(1, 3).value = cost
        sheet_ranges.cell(1, 4).value = newcount
        wb.save('CompanyFile.xlsx')
    else:
        sheet_ranges.cell(i, 3).value = cost
        sheet_ranges.cell(i, 4).value = newcount
        wb.save('CompanyFile.xlsx')
    myportfolio()
    initialwalletvalue()


##    wallet = int(input("Enter your initial investment amount"))
##    def addcompany():
##        newcompany = input("Enter company name")
##        portfolio.append(name)
##        #search(name)
##        return portfolio
##    return portfolio
##def showportfolio():
##    #ALEC IS DOING THIS
##    #ART STUFF TO DISPLAY THE USERS PORTFOLIO
##    #USER SHOULD BE ABLE TO SELECT BEWTWEEN ALL THEIR PORFOLIOS


###RUNNING CODE###


initialportfolio()

master = Tk()
# master.iconbitmap('Cucumber')
master.title('Cucumber')
tab_control = ttk.Notebook(master)
tab1 = ttk.Frame(tab_control)
tab2 = ttk.Frame(tab_control)
tab3 = ttk.Frame(tab_control)
tab4 = ttk.Frame(tab_control)
tab_control.add(tab1, text='Search')
tab_control.add(tab2, text='Portfolio')
tab_control.add(tab3, text='Watchlist')
tab_control.add(tab4, text='My Wallet')
tab_control.grid(column=0,row=0)


lbl1 = Label(tab1, text='')
lbl1.grid(column=3, row=0)
currentP = Label(tab1, text='')
currentPLabel = Label(tab1, text=' Current Price:')
currentPLabel.grid(column=3, row=1)
currentP.grid(column=4, row=1)
OpenP = Label(tab1, text='')
OpenP.grid(column=4, row=2)
OpenPLabel = Label(tab1, text='Opening Price:')
OpenPLabel.grid(column=3, row=2)
DayLow = Label(tab1, text='')
DayLow.grid(column=4, row=3)
DayLowLabel = Label(tab1, text='Daily Low:')
DayLowLabel.grid(column=3, row=3)
DayHigh = Label(tab1, text='')
DayHigh.grid(column=4, row=4)
DayHighLabel = Label(tab1, text='Daily High:')
DayHighLabel.grid(column=3, row=4)
lbl = Label(tab1, text='Please enter the stock ticker or company name')
lbl.grid(column=2, row=1)
txt = Entry(tab1, width=10)
txt.grid(column=2, row=2)
lbl3 = Label(tab3, text='Feature still in development')
lbl3.grid(column=0,row=0)
tab4btn = Button(tab4, text='Update Wallet', command=setwalletvalue)
tab4btn.grid(column=0, row=3)
searchbtn = Button(tab1, text='Search', command=clicked)
searchbtn.grid(column=2, row=3)
portfolioBox = Listbox(tab2)
portfolioBox.grid(column=0,row=0)
for item in portfolio:
    portfolioBox.insert(END,item)
addtoportfoliobtn = Button(tab2, text='Add New Company', command=addtoportfolio)
addtoportfoliobtn.grid(column=2, row=0)
initialwalletvalue()
master.mainloop()

###DIRECTORY CODE###

def directory():
    try:
        userchoice = int(input("\nPlease type the number appropriate to your wanted function, select 0 for all available functions"))
        if userchoice == 0:
            print("Select 1 to find up to date individual stock information")
            print("Select 2 to view your portfolio")
            usermenupick(int(input("Select a function")))
        else:
            usermenupick(userchoice)
    except:
        print("Invalid input")

def usermenupick(value):
    if value == 1:
        initialsearch = str(input("\nWhat is the name of the company?"))
        gosearch(initialsearch)
    elif value == 2:
        myportfolio()
        print("Your current portfolio is:", portfolio)
        userchoice = int(input(" Select 1 - To add a company to your portfolio\n "
                               "Select 2 - To remove a company from your portfolio IN DEVELOPMENT\n"
                               " Select 3 to return to main menu"))
        if userchoice == 1:
            addtoportfolio()

#while 1 == 1:
#    directory()
